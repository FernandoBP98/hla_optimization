# Funciones de utilidad: lectura de datos, representación básica, etc.

import openpyxl


def read_file_excel(filename, size):
    """
    Función que permite la lectura de un fichero .xlsx.

    :param filename: Nombre del fichero que contiene los datos.
    :param size: Número de datos a leer.

    :return: data: Estructura de datos del problema.
    """

    # Inicialización de estructura de datos
    data = {'id': [], 'town': [], 'waste': [], 'longitude': [], 'latitude': [], 'distance': []}
    file = openpyxl.load_workbook(filename)

    # Datos de residuos
    sheet = file['Residuos2014']

    for i in range(2, min(2 + size, sheet.max_row + 1)):
        data['id'].append(i - 2)
        data['town'].append(sheet.cell(i, 1).value)
        data['waste'].append(float(sheet.cell(i, 2).value))

    # Datos de distancia
    sheet = file['Distancias']

    data['distance'] = [-1.0] * int(size + ((size - 2) * (size - 1) / 2) - 1)
    for i in range(2, min(2 + size, sheet.max_row + 1) - 1):
        for j in range(i + 1, min(2 + size, sheet.max_row + 1)):
            ti, tj = i-1, j-1
            data['distance'][int(ti + ((tj - 2) * (tj - 1)) / 2) - 1] = float(sheet.cell(i, j).value)

    # Datos de coordenadas
    sheet = file['Coordenadas']

    for i in range(2, min(2 + size, sheet.max_row + 1)):
        data['latitude'].append(float(sheet.cell(i, 2).value))
        data['longitude'].append(float(sheet.cell(i, 3).value))

    file.close()

    return data

